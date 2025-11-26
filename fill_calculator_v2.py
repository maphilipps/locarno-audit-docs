#!/usr/bin/env python3
"""Fill adesso Calculator 2.01 with Locarno Film Festival Audit Data"""
from openpyxl import load_workbook
from pathlib import Path

def fill_start_sheet(sheet):
    """Fill the Start sheet with project metadata"""
    print("  → Filling Start sheet...")
    sheet["B8"] = "Locarno Film Festival - Drupal CMS 2.0 Relaunch"
    sheet["B10"] = "Locarno Film Festival"
    sheet["B11"] = "adesso Schweiz AG"
    sheet["B12"] = "2025-01"
    sheet["B13"] = "Drupal 11 with Drupal CMS 2.0"

def fill_features_sheet(sheet):
    """Fill Features sheet with 17 Content Types + 35 Paragraphs"""
    print("  → Filling Features sheet (52 features)...")

    content_types = [
        ("F-1", "Homepage", "Website entry point with hero, featured content", "H", 40),
        ("F-2", "Film Program Page", "Dynamic festival program with filters", "H", 60),
        ("F-3", "Film Archive", "Historical film database with search", "H", 50),
        ("F-4", "Film Detail Page", "Individual film presentation", "M", 30),
        ("F-5", "News & Updates", "News listing and detail pages", "M", 25),
        ("F-6", "Event Landing Page", "Festival events and schedule", "M", 35),
        ("F-7", "Press & Media", "Press releases and media kit", "L", 20),
        ("F-8", "Partner & Sponsor", "Partner listings and logos", "L", 15),
        ("F-9", "About Festival", "History, mission, team pages", "M", 25),
        ("F-10", "Industry Section", "Professional accreditation area", "H", 45),
        ("F-11", "Educational Programs", "Workshops and masterclasses", "M", 30),
        ("F-12", "Venue Information", "Theater locations and facilities", "L", 18),
        ("F-13", "Ticketing Integration", "Ticket purchase and booking", "H", 55),
        ("F-14", "User Profile (Keycloak SSO)", "User account and preferences", "H", 40),
        ("F-15", "VOD Streaming Pages", "Sony VOD integration", "H", 50),
        ("F-16", "Search Results", "Site-wide Solr search", "M", 35),
        ("F-17", "FAQ & Help", "User support and documentation", "L", 15),
    ]

    paragraphs = [
        ("F-18", "Hero Carousel", "Full-width image slider with CTAs", "M", 12),
        ("F-19", "Accordion Component", "Expandable content sections", "L", 6),
        ("F-20", "Film Card Grid", "Film teasers in responsive grid", "M", 10),
        ("F-21", "Multi-Column Layout", "Flexible column system", "L", 8),
        ("F-22", "Quote Block", "Highlighted testimonials", "L", 4),
        ("F-23", "Image Gallery", "Lightbox photo gallery", "M", 10),
        ("F-24", "Video Embed", "YouTube/Vimeo integration", "L", 5),
        ("F-25", "CTA Banner", "Call-to-action sections", "L", 6),
        ("F-26", "Event Calendar Widget", "Mini calendar view", "M", 12),
        ("F-27", "News Teaser List", "Latest news overview", "L", 8),
        ("F-28", "Partner Logo Grid", "Sponsor logo display", "L", 6),
        ("F-29", "Statistics Counter", "Animated number counters", "L", 8),
        ("F-30", "Timeline Component", "Event/history timeline", "M", 10),
        ("F-31", "Tab Container", "Tabbed content sections", "L", 8),
        ("F-32", "Download List", "File download links", "L", 5),
        ("F-33", "Social Media Feed", "Social embeds widget", "M", 10),
        ("F-34", "Map Integration", "Google Maps venues", "M", 10),
        ("F-35", "Newsletter Signup", "Email subscription form", "L", 8),
        ("F-36", "Search Box", "Solr search integration", "M", 10),
        ("F-37", "Film Filter Bar", "Program filtering UI", "H", 15),
        ("F-38", "Breadcrumb Navigation", "Site navigation trail", "L", 4),
        ("F-39", "Related Content", "Content recommendations", "M", 8),
        ("F-40", "Share Buttons", "Social sharing links", "L", 4),
        ("F-41", "Author Bio", "Contributor information", "L", 5),
        ("F-42", "Tag Cloud", "Content tagging display", "L", 5),
        ("F-43", "Alert Banner", "Important notices", "L", 5),
        ("F-44", "Progress Indicator", "Visual progress bars", "L", 6),
        ("F-45", "Pricing Table", "Ticket pricing display", "M", 8),
        ("F-46", "Team Member Grid", "Staff directory", "L", 8),
        ("F-47", "Testimonial Slider", "Quote carousel", "M", 10),
        ("F-48", "Icon Feature List", "Feature highlights", "L", 6),
        ("F-49", "Before/After Slider", "Image comparison", "M", 10),
        ("F-50", "Custom HTML Block", "Flexible HTML content", "L", 4),
        ("F-51", "Divider/Spacer", "Visual separators", "L", 2),
        ("F-52", "Anchor Links", "In-page navigation", "L", 4),
    ]

    row = 8
    for fid, name, desc, complexity, hours in content_types + paragraphs:
        sheet[f"A{row}"] = fid
        sheet[f"B{row}"] = name
        sheet[f"C{row}"] = desc
        sheet[f"D{row}"] = complexity
        sheet[f"E{row}"] = hours
        row += 1

def fill_project_tasks_sheet(sheet):
    """Fill Project tasks sheet with 39 tasks across 7 phases"""
    print("  → Filling Project tasks sheet (39 tasks)...")

    tasks = [
        ("T-1", "Setup & Kickoff", "Project kickoff meeting and planning", "PM", 8),
        ("T-2", "Setup & Kickoff", "DDEV local dev environment setup", "Backend", 4),
        ("T-3", "Setup & Kickoff", "Azure infrastructure provisioning", "DevOps", 16),
        ("T-4", "Setup & Kickoff", "CI/CD pipeline configuration", "DevOps", 12),
        ("T-5", "Setup & Kickoff", "Git repository and branching strategy", "Backend", 4),

        ("T-6", "Content Architecture", "Content type modeling (17 types)", "Backend", 40),
        ("T-7", "Content Architecture", "Paragraph component development (35 types)", "Backend", 80),
        ("T-8", "Content Architecture", "Taxonomy vocabularies setup", "Backend", 12),
        ("T-9", "Content Architecture", "Views and listing pages", "Backend", 30),
        ("T-10", "Content Architecture", "Custom fields and field formatters", "Backend", 20),

        ("T-11", "Frontend & Theming", "Mercury Theme setup", "Frontend", 8),
        ("T-12", "Frontend & Theming", "Storybook component development", "Frontend", 60),
        ("T-13", "Frontend & Theming", "Tailwind CSS configuration", "Frontend", 12),
        ("T-14", "Frontend & Theming", "Responsive design implementation", "Frontend", 40),
        ("T-15", "Frontend & Theming", "Accessibility (WCAG 2.1 AA)", "Frontend", 24),

        ("T-16", "Integrations", "Keycloak oAuth2 SSO integration", "Backend", 40),
        ("T-17", "Integrations", "Fiona Festival API integration", "Backend", 60),
        ("T-18", "Integrations", "Sony VOD DRM streaming", "Backend", 50),
        ("T-19", "Integrations", "Mobile App REST API", "Backend", 35),
        ("T-20", "Integrations", "Solr Enterprise Search setup", "Backend", 40),
        ("T-21", "Integrations", "CloudFront CDN configuration", "DevOps", 16),

        ("T-22", "Migration", "Content migration scripts", "Backend", 40),
        ("T-23", "Migration", "Asset migration (images/videos)", "Backend", 20),
        ("T-24", "Migration", "User account migration", "Backend", 16),
        ("T-25", "Migration", "URL redirects (301)", "Backend", 12),
        ("T-26", "Migration", "Data validation and cleanup", "Backend", 20),

        ("T-27", "Testing & QA", "Unit test development", "QA", 30),
        ("T-28", "Testing & QA", "Integration testing", "QA", 25),
        ("T-29", "Testing & QA", "E2E testing (Playwright)", "QA", 30),
        ("T-30", "Testing & QA", "Performance testing", "QA", 16),
        ("T-31", "Testing & QA", "Security audit", "QA", 16),
        ("T-32", "Testing & QA", "UAT coordination", "PM", 12),

        ("T-33", "Deployment", "Staging deployment", "DevOps", 8),
        ("T-34", "Deployment", "Production deployment", "DevOps", 12),
        ("T-35", "Deployment", "DNS and SSL configuration", "DevOps", 6),
        ("T-36", "Deployment", "Monitoring setup (Azure Monitor)", "DevOps", 10),
        ("T-37", "Deployment", "Backup strategy implementation", "DevOps", 8),

        ("T-38", "Project Management", "Weekly status meetings", "PM", 40),
        ("T-39", "Project Management", "Documentation and handover", "PM", 24),
    ]

    row = 8
    for tid, phase, desc, role, hours in tasks:
        sheet[f"A{row}"] = tid
        sheet[f"B{row}"] = phase
        sheet[f"C{row}"] = desc
        sheet[f"D{row}"] = role
        sheet[f"E{row}"] = hours
        row += 1

def fill_project_roles_sheet(sheet):
    """Fill Project roles sheet with 8 Kernpositionen"""
    print("  → Filling Project roles sheet (8 Kernpositionen)...")

    roles = [
        ("R-1", "Project manager", "Senior", "Responsible for overall project delivery, stakeholder communication, timeline and budget management", 0.3),
        ("R-2", "Software architect", "Expert", "Technical lead, system architecture design, code reviews, integration strategy", 0.5),
        ("R-3", "Developer", "Senior", "Backend development #1: Content types, migrations, Keycloak SSO, Fiona API", 1.0),
        ("R-4", "Developer", "Senior", "Backend development #2: Sony VOD, Solr search, mobile API, custom modules", 1.0),
        ("R-5", "Developer", "Middle", "Frontend development: Storybook components, Tailwind CSS, accessibility, responsive design", 1.0),
        ("R-6", "UI/UX Designer", "Middle", "Design system, component wireframes, user flows, accessibility compliance", 0.4),
        ("R-7", "DevOps Engineer", "Middle", "Azure infrastructure, CI/CD pipelines, monitoring, performance optimization", 0.5),
        ("R-8", "Test Engineer", "Middle", "Test automation (Playwright), quality assurance, security testing, UAT coordination", 0.5),
    ]

    row = 8
    for rid, title, level, responsibilities, fte in roles:
        sheet[f"A{row}"] = rid
        sheet[f"B{row}"] = title
        sheet[f"C{row}"] = level
        sheet[f"D{row}"] = responsibilities
        sheet[f"E{row}"] = fte
        row += 1

def fill_project_risks_sheet(sheet):
    """Fill Project risks sheet with 12 identified risks"""
    print("  → Filling Project risks sheet (12 risks)...")

    risks = [
        ("R-1", "Fiona API changes", "External festival API modifications during development", "Medium", "High", "Version locking, abstraction layer, vendor communication"),
        ("R-2", "Keycloak SSO complexity", "OAuth2 integration more complex than anticipated", "Medium", "Medium", "Early prototype, security expert consultation"),
        ("R-3", "Sony VOD DRM requirements", "Streaming DRM implementation challenges", "High", "High", "POC before commitment, vendor support contract"),
        ("R-4", "Solr search performance", "Enterprise search scalability under load", "Medium", "High", "Load testing, caching strategy, index optimization"),
        ("R-5", "Migration data quality", "Legacy content requires significant cleanup", "High", "Medium", "Early migration tests, data validation scripts"),
        ("R-6", "Third-party dependencies", "Critical modules become unmaintained", "Low", "High", "Alternative module research, update monitoring"),
        ("R-7", "Azure cost overruns", "Higher than expected cloud infrastructure costs", "Medium", "Medium", "Cost monitoring, auto-scaling limits, budget alerts"),
        ("R-8", "Browser compatibility", "Cross-browser issues in complex components", "Medium", "Low", "Early browser testing, progressive enhancement"),
        ("R-9", "Performance targets", "Page load times exceed 2-second target", "Medium", "High", "Performance budget, CDN optimization, lazy loading"),
        ("R-10", "Accessibility compliance", "WCAG 2.1 AA compliance gaps discovered late", "Low", "High", "Automated testing from start, accessibility audit"),
        ("R-11", "Scope creep", "Uncontrolled feature additions during development", "High", "Medium", "Change request process, backlog prioritization"),
        ("R-12", "Resource availability", "Key team members unavailable during critical phases", "Medium", "High", "Knowledge sharing, documentation, backup resources"),
    ]

    row = 8
    for rid, name, desc, likelihood, impact, mitigation in risks:
        sheet[f"A{row}"] = rid
        sheet[f"B{row}"] = name
        sheet[f"C{row}"] = desc
        sheet[f"D{row}"] = likelihood
        sheet[f"E{row}"] = impact
        sheet[f"F{row}"] = mitigation
        row += 1

def main():
    template_path = Path("/Users/marc.philipps/Downloads/adesso calculator 2.01 - Default Template CMS 1.xlsm")
    output_path = Path("/Users/marc.philipps/Sites/audit_lucarnofestival.ch/locarno-audit-docs/adesso_calculator_locarno_filled.xlsx")

    if not template_path.exists():
        print(f"❌ Template not found: {template_path}")
        return 1

    print(f"📊 Loading template: {template_path.name}")
    wb = load_workbook(str(template_path))

    print(f"📝 Available sheets: {wb.sheetnames}")

    fill_start_sheet(wb["Start"])
    fill_features_sheet(wb["Features"])
    fill_project_tasks_sheet(wb["Project tasks"])
    fill_project_roles_sheet(wb["Project roles"])
    fill_project_risks_sheet(wb["Project risks"])

    print(f"\n💾 Saving as .xlsx (without macros): {output_path.name}")
    wb.save(str(output_path))
    print(f"✅ Successfully saved to: {output_path}")
    print(f"\n⚠️  NOTE: Saved as .xlsx format (macros removed)")
    print(f"   To recalculate formulas, run: python recalc.py {output_path}")
    return 0

if __name__ == "__main__":
    exit(main())
